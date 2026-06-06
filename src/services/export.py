import openpyxl
from openpyxl.styles import Font, PatternFill

def export_filtered(data, tanggal, jam_mulai, jam_selesai):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    headers = ["No","Unit","Temp","Vib","Noise","Status","Time"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=i).fill = PatternFill("solid", start_color="2563eb")
    
    for i, r in enumerate(data, 2):
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=r['unit'])
        ws.cell(row=i, column=3, value=r['temperature'])
        ws.cell(row=i, column=4, value=r['vibration'])
        ws.cell(row=i, column=5, value=r['noise'])
        ws.cell(row=i, column=6, value=r['status'])
        ws.cell(row=i, column=7, value=str(r['timestamp'])[:16])
    
    for col in 'ABCDEFG':
        ws.column_dimensions[col].width = 15
    
    fname = f"ammonia_export_{tanggal}_{jam_mulai.replace(':','-')}_{jam_selesai.replace(':','-')}.xlsx"
    wb.save(fname)
    return fname


def export_all(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    headers = ["ID","Unit","Temp","Vib","Noise","Status","Time"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=i).fill = PatternFill("solid", start_color="2563eb")
    
    for i, r in enumerate(data, 2):
        ws.cell(row=i, column=1, value=r['id'])
        ws.cell(row=i, column=2, value=r['unit'])
        ws.cell(row=i, column=3, value=r['temperature'])
        ws.cell(row=i, column=4, value=r['vibration'])
        ws.cell(row=i, column=5, value=r['noise'])
        ws.cell(row=i, column=6, value=r['status'])
        ws.cell(row=i, column=7, value=r['timestamp'])
    
    for col in 'ABCDEFG':
        ws.column_dimensions[col].width = 15
    
    from datetime import datetime
    fname = f"ammonia_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    return fname